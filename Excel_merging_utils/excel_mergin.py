# Imports
import os
import pprint
import openpyxl

path = input('Type your folder:')
current_dir = os.listdir(path)
pp = pprint.PrettyPrinter()
pp.pprint(current_dir)

# Create new file
merging_wb = openpyxl.load_workbook(path+'\\'+'bunch.xlsx')

# Working on files
for file in current_dir:							
  print('\n'*2)
  print(file)
  # Work with file if it's not as below
  if file != 'bunch.xlsx':
    filename = path+'\\'+file
    print(filename)
    print(f'Opening {file}....')
    xls = openpyxl.load_workbook(filename)
    wsxls = xls.active
    # Create new sheet and name it with as file name has
    wsmxls = merging_wb.create_sheet(wsxls.title)
    
    for row in wsxls:
      for cell in row:
        # Copiing operation: from cell to cell
        wsmxls[cell.coordinate].value = cell.value

    merging_wb.save(path+'\\'+'bunch.xlsx')

print('Succesfully copied!')